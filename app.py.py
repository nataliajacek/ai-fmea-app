import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openai import OpenAI
import os
import json
import pdfplumber

client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

st.title("AI-Assisted FMEA Generator")

# -----------------------------
# PROJECT INFO
# -----------------------------

project = st.text_input("Project")
user_name = st.text_input("User Name")
version = st.text_input("Version")
object_name = st.text_input("Object Name")

st.subheader("Product Requirements / Functions")

inputs = st.text_area("Enter one requirement per line")

# -----------------------------
# FILE UPLOAD
# -----------------------------

st.subheader("Upload Supporting Files (Optional)")

uploaded_files = st.file_uploader(
    "Upload CAD screenshots, PDFs, requirement docs, diagrams",
    accept_multiple_files=True
)

def extract_file_text(files):

    text_data = ""

    for file in files:

        if file.type == "application/pdf":

            with pdfplumber.open(file) as pdf:
                for page in pdf.pages:
                    text_data += page.extract_text() + "\n"

        else:
            try:
                text_data += file.read().decode("utf-8")
            except:
                pass

    return text_data

file_context = ""

if uploaded_files:
    file_context = extract_file_text(uploaded_files)

# -----------------------------
# TEST MATRIX
# -----------------------------

test_columns = [
"INVESTIGATION & TESTING",
"VENDOR - PART",
"DESIGN CHANGE",
"DIM & WORST CASE",
"SIMULATION",
"CHARACTERIZE",
"CPPP",
"DIAGNOSTICS",
"FUNCTIONALITY",
"OOBE & INSTALL",
"SYSTEM TEST",
"SIT E2E APP",
"HALT",
"ALT",
"ROBUSTNESS",
"REGS EMC",
"REGSSAFETY",
"USABILITY",
"SW-FW TESTS",
"MFG TESTS",
"MAINTENANCE",
"SERVICEABILITY"
]

# -----------------------------
# AI GENERATION
# -----------------------------

def generate_fmea(object_name, inputs_text, file_context):

    requirements = [r.strip() for r in inputs_text.split("\n") if r.strip()]
    rows = []

    context = f"""
Product: {object_name}

User requirements:
{inputs_text}

Additional product documentation:
{file_context}
"""

    for req in requirements:

        prompt = f"""
You are a senior reliability engineer performing FMEA.

Product information:
{context}

Function / Requirement:
{req}

Generate at least 10 realistic failure modes.

For each failure return:

Failure Scenario
Part
Failure Mode
End Effects
Causes (2-3)
Current Design Controls
Recommended Actions (2-3)

Owner (choose from: Mechanical Engineering, Electrical Engineering,
Reliability Engineering, Quality Engineering, Manufacturing,
Firmware/Software, UX/Human Factors)

Execution Phase (Concept, Design, Prototype, Validation, Production, Field)

Severity (1-10)
Occurrence (1-10)
Detectability (1-10)

Estimated RPN2 after mitigation.

Also recommend relevant test strategies from this list:

INVESTIGATION & TESTING
VENDOR - PART
DESIGN CHANGE
DIM & WORST CASE
SIMULATION
CHARACTERIZE
CPPP
DIAGNOSTICS
FUNCTIONALITY
OOBE & INSTALL
SYSTEM TEST
SIT E2E APP
HALT
ALT
ROBUSTNESS
REGS EMC
REGSSAFETY
USABILITY
SW-FW TESTS
MFG TESTS
MAINTENANCE
SERVICEABILITY

Return them in "tests".

Also include 1-2 reference links explaining the failure risk.

Return ONLY JSON.
"""

        response = client.chat.completions.create(
            model="gpt-4.1-mini",
            messages=[{"role":"user","content":prompt}],
            temperature=0.2
        )

        text = response.choices[0].message.content

        try:
            failures = json.loads(text)
        except:
            continue

        for f in failures:

            S = f.get("Severity",5)
            O = f.get("Occurrence",5)
            D = f.get("Detectability",5)

            cost = 1
            RPN = S*O*D

            row = {

                "Failure Scenario": f.get("Failure Scenario",""),
                "Function": req,
                "Part": f.get("Part",""),
                "Failure Mode": f.get("Failure Mode",""),
                "End Effects of Failure": f.get("End Effects",""),
                "Causes": ", ".join(f.get("Causes",[])),
                "Current Design Controls": f.get("Controls",""),

                "Severity (S)": S,
                "Occurrence (O)": O,
                "Detectability (D)": D,
                "RPN": RPN,

                "Cost": cost,
                "Priority": RPN*cost,

                "Recommended Actions": ", ".join(f.get("Actions",[])),
                "Owner": f.get("Owner",""),
                "Execution Phase": f.get("Execution Phase",""),

                "Reference Links": ", ".join(f.get("References",[])),
                "RPN2 (Post-Action)": f.get("RPN2","")
            }

            recommended_tests = f.get("tests",[])

            for col in test_columns:
                row[col] = "X" if col in recommended_tests else ""

            rows.append(row)

    return pd.DataFrame(rows)

# -----------------------------
# GENERATE BUTTON
# -----------------------------

if st.button("Generate FMEA"):

    if object_name.strip()=="":
        st.warning("Enter Object Name")
    else:

        df = generate_fmea(object_name,inputs,file_context)

        if df.empty:
            st.warning("Enter at least one requirement")
        else:
            st.session_state.df = df

# -----------------------------
# EDITABLE TABLE
# -----------------------------

if "df" in st.session_state:

    st.subheader("Editable FMEA Table")

    edited_df = st.data_editor(
        st.session_state.df,
        use_container_width=True
    )

    edited_df["RPN"] = (
        edited_df["Severity (S)"] *
        edited_df["Occurrence (O)"] *
        edited_df["Detectability (D)"]
    )

    edited_df["Priority"] = edited_df["RPN"] * edited_df["Cost"]

    st.session_state.df = edited_df

    st.subheader("Updated Calculations")
    st.dataframe(edited_df)

    # -----------------------------
    # EXCEL EXPORT
    # -----------------------------

    wb = Workbook()
    ws = wb.active
    ws.title = "FMEA"

    headers = list(edited_df.columns)
    ws.append(headers)

    for _,row in edited_df.iterrows():
        ws.append(list(row))

    col_map = {name:idx+1 for idx,name in enumerate(headers)}

    for i in range(2,len(edited_df)+2):

        S = col_map["Severity (S)"]
        O = col_map["Occurrence (O)"]
        D = col_map["Detectability (D)"]
        RPN = col_map["RPN"]
        COST = col_map["Cost"]
        PRIORITY = col_map["Priority"]

        ws.cell(row=i,column=RPN).value = f"={chr(64+S)}{i}*{chr(64+O)}{i}*{chr(64+D)}{i}"
        ws.cell(row=i,column=PRIORITY).value = f"={chr(64+RPN)}{i}*{chr(64+COST)}{i}"

    output = BytesIO()
    wb.save(output)

    st.download_button(
        "Download Excel File",
        output.getvalue(),
        file_name=f"FMEA_{object_name}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
